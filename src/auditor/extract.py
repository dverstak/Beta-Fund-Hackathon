"""Vision + spreadsheet extraction -> normalized LineItem list.

Receipts and 1099-NEC scans go to GMI Cloud's multimodal model. Spreadsheets
are parsed locally (pandas if available, else stdlib csv) -- no tokens spent.
"""
from __future__ import annotations

import csv
from pathlib import Path

from .gmi_client import GMIClient
from .models import LineItem

RECEIPT_INSTRUCTION = (
    "This is a business expense receipt (it may be skewed, low quality, or "
    "handwritten). Extract the merchant, date, total amount, and a short "
    "description of what was purchased. Infer the most likely business purpose."
)
RECEIPT_SCHEMA = (
    '{"vendor": str, "date": "YYYY-MM-DD", "total": number, '
    '"description": str, "payment_method": str|null, "currency": str}'
)

FORM_INSTRUCTION = (
    "This is a 1099-NEC (Nonemployee Compensation) form. Extract the payer "
    "name, the nonemployee compensation amount (Box 1), and federal tax "
    "withheld (Box 4) if present."
)
FORM_SCHEMA = (
    '{"payer": str, "box1_nonemployee_comp": number, '
    '"box4_federal_tax_withheld": number, "tax_year": number}'
)


def extract_image(client: GMIClient, path: Path, profile: str) -> list[LineItem]:
    if profile == "form_1099nec":
        d = client.extract_from_image(profile, path, FORM_INSTRUCTION, FORM_SCHEMA)
        amt = _num(d.get("box1_nonemployee_comp"))
        return [LineItem(
            source=path.name, profile=profile, kind="income",
            vendor=d.get("payer"), date=_year_date(d.get("tax_year")),
            description="1099-NEC nonemployee compensation (Box 1)",
            amount=amt, raw=d,
        )]
    # receipt
    d = client.extract_from_image(profile, path, RECEIPT_INSTRUCTION, RECEIPT_SCHEMA)
    return [LineItem(
        source=path.name, profile="receipt", kind="expense",
        vendor=d.get("vendor"), date=d.get("date"),
        description=d.get("description"),
        amount=_num(d.get("total")),
        payment_method=d.get("payment_method"), raw=d,
    )]


def extract_spreadsheet(path: Path) -> list[LineItem]:
    rows = _read_rows(path)
    items: list[LineItem] = []
    for r in rows:
        lower = {k.lower().strip(): v for k, v in r.items() if k}
        amount = _num(_pick(lower, "amount", "total", "debit", "cost", "price"))
        if amount == 0.0:
            continue
        kind = "income" if _pick(lower, "type", "kind").lower() in (
            "income", "revenue", "deposit") else "expense"
        items.append(LineItem(
            source=path.name, profile="spreadsheet", kind=kind,
            date=_norm_date(_pick(lower, "date", "transaction date")),
            vendor=_pick(lower, "vendor", "merchant", "payee", "name"),
            description=_pick(lower, "description", "memo", "details", "category"),
            amount=amount,
            payment_method=_pick(lower, "payment", "method", "account") or None,
            raw=r,
        ))
    return items


# ---- helpers ----
def _read_rows(path: Path) -> list[dict]:
    ext = path.suffix.lower()
    if ext in (".csv", ".tsv"):
        delim = "\t" if ext == ".tsv" else ","
        with path.open(newline="", encoding="utf-8-sig") as f:
            return list(csv.DictReader(f, delimiter=delim))
    if ext in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise RuntimeError("openpyxl required for .xlsx ingest") from e
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(h) if h is not None else "" for h in rows[0]]
        return [dict(zip(header, r)) for r in rows[1:]]
    return []


def _pick(d: dict, *keys: str) -> str:
    for k in keys:
        if k in d and d[k] not in (None, ""):
            return str(d[k]).strip()
    return ""


def _num(v) -> float:
    if v is None or v == "":
        return 0.0
    s = str(v).replace("$", "").replace(",", "").strip()
    s = s.replace("(", "-").replace(")", "")
    try:
        return round(float(s), 2)
    except ValueError:
        return 0.0


def _norm_date(s: str) -> str | None:
    return s or None


def _year_date(year) -> str | None:
    try:
        return f"{int(year)}-12-31"
    except (TypeError, ValueError):
        return None
